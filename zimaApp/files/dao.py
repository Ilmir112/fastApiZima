
import mimetypes
from datetime import datetime, timedelta
import re
from zoneinfo import ZoneInfo

from bson import ObjectId
from fastapi import UploadFile, Request

from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from openpyxl.reader.excel import load_workbook

from zimaApp.database import ImageMongoDB  # предполагаю, что это модель/схема
from zimaApp.logger import logger

from zimaApp.repair_data.schemas import RepairDataCreate


class MongoFile:
    @classmethod
    async def upload_file(cls, request: Request, itemId: str, file: UploadFile):
        try:
            db = request.app.state.mongo_client["files"]
            fs_bucket = AsyncIOMotorGridFSBucket(db)
            filename = file.filename
            contents = await file.read()

            # Определяем MIME тип
            mime_type, _ = mimetypes.guess_type(filename)
            content_type = mime_type or "application/octet-stream"

            # Загружаем содержимое в GridFS
            gridfs_id = await fs_bucket.upload_from_stream(
                filename,
                contents,
                metadata={
                    "contentType": content_type,
                    "field_id": itemId,
                    "upload_date": datetime.utcnow(),
                    "size": len(contents),
                }
            )

            # Создайте экземпляр модели
            document = ImageMongoDB(
                name=filename,
                file_id=str(gridfs_id)  # добавьте ID файла в GridFS
            )

            # Вставляем документ в коллекцию (предположим, что у вас есть модель)
            result = await ImageMongoDB.insert_one(document)
            file_id_str = str(result.id)

            # Формируем ссылку на файл (может быть API для скачивания по ID)
            file_url = f"/files/{file_id_str}"

            return file_url, file_id_str, filename
        except Exception as e:
            logger.error(e)

    @classmethod
    async def get_file_from_mongo(cls, request: Request, file_id: str):
        try:
            db = request.app.state.mongo_client["files"]
            fs_bucket = AsyncIOMotorGridFSBucket(db)
            # Получаем документ по ID
            doc = await ImageMongoDB.find_one({"_id": ObjectId(file_id)})
            if not doc:
                raise Exception("Файл не найден")

            # Получаем file_id из документа
            gridfs_id = doc.file_id  # предполагается, что это поле есть и правильно заполнено
            if not gridfs_id:
                raise Exception("ID файла в GridFS отсутствует")

            # Если gridfs_id — строка, преобразуйте в ObjectId
            if isinstance(gridfs_id, str):
                gridfs_id = ObjectId(gridfs_id)

            # Открываем поток для скачивания
            stream = await fs_bucket.open_download_stream(gridfs_id)
            return stream  # Можно вернуть поток или прочитать полностью
        except Exception as e:
            logger.error(e)

    @classmethod
    async def delete_file_from_mongo(cls, request: Request, file_id: str):
        db = request.app.state.mongo_client["files"]
        fs_bucket = AsyncIOMotorGridFSBucket(db)
        fs_bucket.delete(ObjectId(file_id))
        return True


class ExcelRead:
    def __init__(self, df):
        self.df = df

    def find_pars(self):
        # Проверка наличия колонок
        required_columns = ['Дата', 'Работы', 'Примечание']
        for col in required_columns:
            if col not in self.df.columns:
                print(f"Отсутствует обязательная колонка: {col}")
                return None  # Или выбросить исключение

        # Проверка, что есть хотя бы одна строка
        if self.df.empty:
            print("Данных нет.")
            return None
        skv_number = None
        region = None
        mesto_matches = []
        for row in self.df.itertuples():
            # Объединение всех ячеек первой строки в одну строку для поиска
            row_text = ' '.join(str(cell) for cell in row)
            if '> Начало ремонта' in row_text:
                row_begin = row_text.split(';')[0]
                skv_pattern = r'№(\d+)...'

                # Регулярное выражение для даты (день.месяц.год)
                date_pattern = r'(\d{2}\.\d{2}\.\d{4})'

                # Регулярное выражение для времени начала (первое время в диапазоне)
                time_pattern = r'(\d+:\d{2}) - (\d+:\d{2})'

                # Поиск даты
                date_match = re.search(date_pattern, row_begin)
                date_value = date_match.group(1) if date_match else None

                # Поиск времени начала (первое время в диапазоне)
                time_match = re.search(time_pattern, row_begin)
                start_time = time_match.group(1) if time_match else None

                # Поиск номера скважины
                skv_match = re.search(skv_pattern, row_begin)
                skv_number = skv_match.group(1) if skv_match else None
                if ' КР)' in row_begin:
                    region = "КГМ"
                elif ' ТР)' in row_begin:
                    region = "ТГМ"
                elif ' ИР)' in row_begin:
                    region = "ИГМ"
                elif ' АР)' in row_begin:
                    region = "АГМ"
                elif ' ЧР)' in row_begin:
                    region = "ЧГМ"
            mesto_matches = re.findall(r'на скв\.?\s*№\s*[\w\-]+(?:\s+)([\w\-]+)', row_text, re.IGNORECASE)
            if ('переезд' in row_text.lower() or 'перестанов' in row_text.lower()) and len(mesto_matches) != 0:
                break

        # if skv_number is None or mesto_matches[0] is None:
        #     return

        return skv_number, mesto_matches, region, date_value + " " + start_time

    @staticmethod
    def extract_datetimes(row):
        # Обработка строки с датой
        date_str = row[1].split('\n')[0]  # '03.08.2025'
        time_range = row[1].split('\n')[1]  # '18:00-22:00'

        # Парсим дату
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')

        # Парсим время начала и конца интервала
        start_time_str, end_time_str = time_range.split('-')
        start_time = datetime.strptime(start_time_str.strip(), '%H:%M').time()
        end_time = datetime.strptime(end_time_str.strip(), '%H:%M').time()


        # Создайте datetime без временной зоны
        start_datetime_naive = datetime.combine(date_obj, start_time)

        # Затем установите временную зону
        start_datetime = start_datetime_naive.replace(tzinfo=ZoneInfo("Asia/Yekaterinburg"))
        end_datetime = datetime.combine(date_obj.date(), end_time)

        return (start_datetime, row[2])

    @staticmethod
    def parse_datetime(value):
        if isinstance(value, datetime):
            return value
        elif isinstance(value, str):
            # Попытка распарсить строку
            try:
                return datetime.strptime(value, "%d.%m.%Y %H:%M")
            except ValueError:
                try:
                    return datetime.strptime(value, "%d.%m.%Y")
                except ValueError:
                    return None
        return None

    @classmethod
    async def read_repairs_well_excel(cls, filename):
        from zimaApp.repair_data.router import add_repair_data

        workbook = load_workbook(filename)
        sheet = workbook['Каталог_ремонтов']
        if sheet:

            # Получение данных из Excel и запись их в базу данных
            for index_row, row in enumerate(sheet.iter_rows(min_row=1, max_row=7, max_col=38, values_only=True)):
                for col, value in enumerate(row):

                    if not value is None:
                        if "Каталог ремонтов ООО «Башнефть-Добыча»" in str(value):
                            date_data_str = value.split(" ")[-1]
                            date_data = datetime.strptime(date_data_str, "%d.%m.%Y")
                        # print(value)
                        if 'подрядчик' == str(value).lower():
                            row_index = index_row + 5
                            contractor_index = col
                        elif 'Бригада' in str(value):
                            number_brigade_index = col
                        elif 'Площадь' in str(value):
                            well_area = col
                        elif '№ скв.' in str(value):
                            well_number = col
                        elif 'Фактические дата и время начала' in str(value):
                            begin_index = col
                        elif 'Фактические дата и время окончания' in str(value):
                            finish_index = col
                        elif 'Категория ремонта' in str(value):
                            category_well_index = col
                        elif 'Уникальный код ремонта' in str(value):
                            unicum_index = col
                        elif 'Вид ремонта' in str(value):
                            type_repair_index = col
                        elif 'Продолжительность ремонта,  час' in str(value):
                            duration_index = col
                        elif 'Куст' in str(value):
                            bush_index = col

        for index_row, row in enumerate(sheet.iter_rows(min_row=row_index, values_only=True)):
            if 'Ойл-С' in str(row[contractor_index]):
                try:
                    begin_dt = cls.parse_datetime(row[begin_index])
                    begin_dt = cls.round_to_nearest_30_minutes(begin_dt)

                    if row[finish_index]:
                        finish_dt = cls.parse_datetime(row[finish_index])
                        finish_dt = cls.round_to_nearest_30_minutes(finish_dt)
                    else:
                        finish_dt = None

                    # Проверка по дате для finish_time (если нужно)
                    if finish_dt and row[finish_index].date() != date_data.date():
                        finish_dt = finish_dt  # оставить как есть или дополнительно обработать
                    else:
                        finish_dt = None
                    data = RepairDataCreate(
                        contractor=row[contractor_index],
                        brigade_number=row[number_brigade_index],
                        well_area=row[well_area],
                        well_number=row[well_number],
                        begin_time=begin_dt,
                        finish_time=finish_dt,
                        category_repair=str(row[category_well_index]) if row[category_well_index] else None,
                        repair_code=str(row[unicum_index]),
                        type_repair=str(row[type_repair_index]) if row[type_repair_index] else None,
                        duration_repair=row[duration_index],
                        bush=str(row[bush_index])
                    )
                    result = await add_repair_data(data)
                except Exception as e:
                    print(e)

        return result

    @staticmethod
    def round_to_nearest_30_minutes(dt):
        # Получаем количество минут
        minutes = dt.minute
        # Определяем, к какой границе ближе: 0 или 30
        if minutes % 30 >= 15:
            # Округляем вверх до следующего 30-минутного интервала
            delta_minutes = 30 - (minutes % 30)
            dt = dt + timedelta(minutes=delta_minutes)
        else:
            # Округляем вниз
            delta_minutes = minutes % 30
            dt = dt - timedelta(minutes=delta_minutes)
        # Обнуляем секунды и микросекунды
        return dt.replace(second=0, microsecond=0)